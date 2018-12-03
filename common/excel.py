# coding: utf-8
import logging
import re
from itertools import chain, zip_longest

from django.db.models.fields.files import FieldFile
from openpyxl.worksheet.datavalidation import DataValidation

from common.models import MetaData
from common.utils import decimal, parsedate, str_to_bool, json_encode, patch_settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.translation import ugettext_lazy as _
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Logging
logger = logging.getLogger(__name__)

# Types de données
TYPES = {
    'AutoField': _('Identifiant'),
    'BigIntegerField': _('Nombre entier'),
    'BinaryField': _('Données binaires'),
    'BooleanField': _('Booléen'),
    'CharField': _('Chaîne de caractères'),
    'CommaSeparatedIntegerField': _('Nombres entiers séparés par des virgules'),
    'DateField': _('Date'),
    'DateTimeField': _('Date & heure'),
    'DecimalField': _('Nombre décimal'),
    'DurationField': _("Durée"),
    'EmailField': _('E-mail'),
    'FileField': _('Fichier'),
    'FilePathField': _('Chemin de fichier'),
    'FloatField': _('Nombre flottant'),
    'ImageField': _('Image'),
    'IntegerField': _('Nombre entier'),
    'IPAddressField': _('Adresse IP'),
    'GenericIPAddressField': _('Adresse IP générique'),
    'NullBooleanField': _('Booléen à 3 états'),
    'PositiveIntegerField': _('Nombre entier positif'),
    'PositiveSmallIntegerField': _('Nombre entier positif'),
    'SlugField': _('Slug'),
    'SmallIntegerField': _('Nombre entier'),
    'TextField': _('Texte'),
    'TimeField': _('Heure'),
    'URLField': _('URL'),
    'UUIDField': _('UUID'),
    'ForeignKey': _('Référence'),
    'ManyToManyField': _('Références multiples'),
    'OneToOneField': _('Référence'),
}
CELL_OFFSET = 3

METADATA_NAME = _("Métadonnées")
DROPDOWN_NAME = _("Listes")


class ImportExport(object):
    def __init__(self, models, log=logger, force=False, clean=True, non_editables=False):
        """
        :param models: Liste des modèles à exporter/importer
        :param log: Logger
        :param force: Force l'insertion même en cas d'erreur
        :param clean: Exécute les tests de validation
        :param non_editables: Exporte les données non éditables
        """
        self.models = models
        self.log = log
        self.force = force
        self.clean = clean
        self.non_editables = non_editables

    @transaction.atomic
    def importer(self, file):
        """
        Importe les données d'un document Excel de tarification
        :param file: Chemin vers le document Excel
        :return: Cache
        """
        cache = {}
        metadata = {}

        workbook = load_workbook(filename=file, read_only=True, data_only=True)
        # Récupération de toutes les feuilles par nom
        worksheets = {}
        for worksheet in workbook.worksheets:
            worksheets[worksheet.title.lower()] = worksheet

        # Si elle existe, nous traitons la feuille des métadonnées
        metadata_sheet_name = str(METADATA_NAME)
        if metadata_sheet_name in worksheets:
            worksheet = worksheets.get(metadata_sheet_name)
            headers = {}
            title = True
            for row_number, row in enumerate(worksheet.iter_rows()):
                code_meta = ''
                line = []
                for col_number, cell in enumerate(row):
                    value = cell.value
                    if isinstance(value, str):
                        value = value.strip()
                    if value is None or not str(value).strip():
                        continue
                    # Si c'est la ligne des titres, on ne récupère que les données liées aux colonnes
                    if title:
                        value = value.lower()
                        headers[col_number] = value
                        continue
                    field = headers[col_number]
                    if field == 'code':
                        if value not in metadata:
                            metadata[value] = []
                        code_meta = value
                        continue
                    line.append(value)
                # Si c'est la ligne des titres, on n'enregistre aucune donnée
                if title:
                    title = False
                    continue
                metadata[code_meta].append(line)

        done = []
        for model in self.models:
            code_field = getattr(model, '_code_field', 'id')
            # Retrait des espaces et des caractères superflus
            model_name = re.sub(r'[^\w]+', ' ', str(model._meta.verbose_name).lower())
            # Récupération de la feuille correspondante au modèle
            if model_name not in worksheets:
                self.log.warning(_(
                    "La feuille correspondant au modèle '{model_name}' "
                    "n'a pu être trouvée dans le fichier.").format(model_name=model_name))
                continue
            worksheet = worksheets.get(model_name)
            # Récupération des champs du modèle
            fields = {}
            for field in chain(model._meta.fields, model._meta.many_to_many):
                if field.name != code_field and (field.auto_created or not (field.editable or self.non_editables)):
                    continue
                field.m2m = field in model._meta.many_to_many
                fields[str(field.verbose_name).lower()] = field
            # Parcours des lignes de la feuille
            self.delayed_models = []
            headers = {}
            title = True
            for row_number, row in enumerate(worksheet.iter_rows()):
                instance = model()
                current_metadata = {}
                delayed = False
                m2m = {}
                fks = {}
                # Parcours des cellules de la ligne
                has_data = False
                for col_number, cell in enumerate(row):
                    # Récupération de la valeur de la cellule, ignorée si vide
                    value = cell.value
                    if isinstance(value, str):
                        value = value.strip()
                    if value is None or not str(value).strip():
                        continue
                    # Si c'est la ligne des titres, on ne récupère que les données liées aux colonnes
                    if title:
                        value = value.lower()
                        if value in fields:
                            headers[col_number] = fields[value]
                        continue
                    # Si la colonne n'est pas référencée comme un champ connu, elle est ignorée
                    if col_number not in headers:
                        continue
                    field = headers[col_number]
                    # Gestion des types spécifiques mal gérés par Excel
                    type = field.get_internal_type()
                    if field.m2m:
                        if field.related_model == model:
                            delayed = True
                        value = [v.strip() for v in value.split(',')]
                        m2m[field.name] = (field.related_model, value)
                        has_data = True
                        continue
                    elif field.remote_field is not None and field.related_model is MetaData:
                        current_metadata = dict(metadata.get(value, []))
                        continue
                    elif field.remote_field:
                        if field.related_model == model:
                            delayed = True
                        fks[field.name] = (field.related_model, value)
                        has_data = True
                        continue
                    elif field.choices:
                        choices = {str(value): str(key) for key, value in field.flatchoices}
                        if hasattr(field, 'max_choices'):  # MultiSelectField
                            value = [choices[val] for val in choices.keys() if val in value]
                        else:
                            value = choices[value]
                    elif type in ['DateField', 'DateTimeField']:
                        value = parsedate(value, dayfirst=True)
                    elif type == 'DecimalField':
                        value = decimal(value, precision=20)
                    elif type == 'BooleanField':
                        value = str_to_bool(value)
                    has_data = True
                    # Récupération des données existantes
                    if field.name == code_field and field.unique:
                        existing = model.objects.filter(**{code_field: value})
                        if existing.count() == 1:
                            instance = existing.first()
                    # Modification des propriétés du modèle
                    setattr(instance, field.name, value)
                # Si c'est la ligne des titres, on n'enregistre aucune donnée
                if title:
                    title = False
                    continue
                # Si la ligne est vide, on passe à la suivante
                if not has_data:
                    continue
                # Mise en cache de l'instance  courante
                code = getattr(instance, code_field, id(instance))
                if model not in cache:
                    cache[model] = {}
                cache[model][code] = instance
                # Enregistrement immédiat (si possible)
                if delayed:
                    self.delayed_models.append((instance, fks, m2m, current_metadata))
                    continue
                self._save_instance(instance, metadata=current_metadata, cache=cache, fks=fks, m2m=m2m)
            # Enregistrement différé
            for instance, fks, m2m, current_metadata in self.delayed_models:
                self._save_instance(instance, metadata=current_metadata, cache=cache, fks=fks, m2m=m2m)
            # Intégration terminée
            done.append(model)
        return cache

    @transaction.atomic
    def exporter(self, file):
        """
        Exporte les données de tarification dans un document Excel
        :param file: Chemin vers le document Excel
        :return: Rien
        """
        workbook = Workbook()
        self.dropdowns = {}
        # Style des titres
        self.title_font = Font(bold=True)
        self.metadata = {}

        # Feuille d'aide sur les données
        worksheet = workbook.active
        worksheet.title = str(_("Informations"))
        titles = [_("modèle"), _("champ"), _("type"), _("description")]
        widths = {}
        for column, title in enumerate(titles, start=1):
            cell = worksheet.cell(row=1, column=column)
            cell.value = str(title)
            cell.font = self.title_font
            column_letter = get_column_letter(column)
            widths[column_letter] = len(str(cell.value)) + CELL_OFFSET
        row = 2
        for model in self.models:
            meta = model._meta
            for field in meta.fields + meta.many_to_many:
                if field.auto_created or not (field.editable or self.non_editables):
                    continue
                datas = [
                    meta.verbose_name.capitalize(),
                    field.verbose_name,
                    TYPES[field.get_internal_type()],
                    field.help_text]
                if field.choices:
                    self.dropdowns[model, field.name] = [str(value) for key, value in field.flatchoices]
                for column, data in enumerate(datas, start=1):
                    cell = worksheet.cell(row=row, column=column)
                    cell.value = str(data)
                    column_letter = get_column_letter(column)
                    widths[column_letter] = max(widths[column_letter], len(str(data)) + CELL_OFFSET)
                row += 1
        # Redimensionne les colonnes
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width

        # Listes déroulantes
        worksheet = workbook.create_sheet(title=str(DROPDOWN_NAME))
        worksheet.sheet_state = 'hidden'
        for row in zip_longest(*self.dropdowns.values(), fillvalue=None):
            worksheet.append(row)
        for index, key in enumerate(self.dropdowns.keys(), start=1):
            column = get_column_letter(index)
            self.dropdowns[key] = DataValidation(
                type='list', formula1='={}!${}:${}'.format(DROPDOWN_NAME, column, column))

        # Feuille par modèle
        for model in self.models:
            self._write_model(workbook, model)

        # Export des métadatas
        worksheet = workbook.create_sheet(title=str(METADATA_NAME))
        fields = [('code', 'Code'), ('cle', 'Clé'), ('valeur', 'Valeur')]
        for column, (field_code, field_name) in enumerate(fields, start=1):
            cell = worksheet.cell(row=1, column=column)
            cell.value = field_name
            cell.font = self.title_font
            column_letter = get_column_letter(column)
            widths[column_letter] = len(str(cell.value)) + CELL_OFFSET
        # On construit la feuille des métadonnées ligne par ligne en bouclant sur notre dictionnaire de métadonnées
        row = 2
        for id, liste_tuple_meta in self.metadata.items():
            for key, value in liste_tuple_meta:
                # La colonne 1 correspond au code
                cell = worksheet.cell(row=row, column=1)
                cell.value = id
                # La colonne 2 correspond à la clé
                cell = worksheet.cell(row=row, column=2)
                cell.value = key
                # La colonne 3 correspond à la valeur
                cell = worksheet.cell(row=row, column=3)
                try:
                    cell.value = value
                except Exception:
                    cell.value = json_encode(value)
                row += 1

        workbook.save(file)

    def _save_instance(self, instance, metadata, cache, fks=None, m2m=None):
        """
        Enregistre l'instance en base de données
        :param instance: Instance à sauvegarder
        :param metadata: Metadonnées liées à l'instance
        :param cache: Autres instances en cache (optimisation)
        :param fks: Liste des clés étrangères
        :param m2m: Listes de relations de type many-to-many
        :return: Instance
        """
        # Enregistrement des clés étrangères
        try:
            for field_name, (related, value) in fks.items():
                code_field = getattr(related, '_code_field', 'id')
                fk = cache.get(related, {}).get(value, related.objects.get(**{code_field: value}))
                setattr(instance, field_name, fk)
        except Exception:
            if self.delayed_models:
                # On va chercher l'instance parent et l'enregistrer en amont
                for index, (parent, _fks, _m2m, _metadata) in enumerate(self.delayed_models):
                    if parent.code == value:
                        self._save_instance(parent, metadata=_metadata, cache=cache, fks=_fks, m2m=_m2m)
                        self.delayed_models.pop(index)
                        break
                else:
                    logger.error(_("Impossible de récupérer la valeur de clé étrangère "
                                   "correspondant à [{}] pour le champ [{}] de [{}]").format(
                        value, field_name, instance._meta.verbose_name))
                    raise
            else:
                logger.error(_("Impossible de récupérer la valeur de clé étrangère "
                               "correspondant à [{}] pour le champ [{}] de [{}]").format(
                    value, field_name, instance._meta.verbose_name))
                raise
        # Tests de validation et enregistrement de l'instance
        try:
            code_field = getattr(instance, '_code_field', 'id')
            if not getattr(instance, code_field, None):
                instance.validate_unique()
            if self.clean:
                instance.clean()
            with patch_settings(IGNORE_LOG=True):
                instance.save()
        except ValidationError as errors:
            for field, errors in errors.message_dict.items():
                for error in errors:
                    if field == '__all__':
                        self.log.warning(error)
                    else:
                        self.log.warning('[{}] {}'.format(field, error))
            if not self.force:
                raise
        # Enregistrement des métadonnées (possible qu'après l'enregistrement en base)
        try:
            for key, value in metadata.items():
                instance.set_metadata(key, value)
        except Exception:
            logger.error(_("Impossible d'ajouter la métadata [{},{}] pour l'instance '[{}]'").format(
                key, value, instance._meta.verbose_name))
            raise
        # Enregistrement des many-to-many sur l'instance (possible qu'après l'enregistrement en base)
        try:
            for field_name, (related, values) in m2m.items():
                code_field = getattr(related, '_code_field', 'id')
                m2ms = [cache.get(related, {}).get(value, related.objects.get(**{code_field: value})) for value in values]
                getattr(instance, field_name).set(m2ms)
        except Exception:
            logger.error(_("Impossible de récupérer les valeurs de relation "
                           "correspondantes à [{}] pour le champ [{}] de [{}]").format(
                ', '.join(str(v) for v in values), field_name, instance._meta.verbose_name))
            raise
        self.log.info('{} : {}'.format(instance._meta.verbose_name.capitalize(), instance))
        return instance

    def _write_model(self, workbook, model):
        """
        Ecrit le modèle dans un document Excel ouvert
        :param workbook: Document Excel
        :param model: Modèle
        :return: Rien
        """
        meta = model._meta
        code_field = getattr(model, '_code_field', 'id')
        worksheet = workbook.create_sheet(title=re.sub(r'[^\w]+', ' ', str(meta.verbose_name).capitalize()))
        widths = {}
        dropdowns = {}
        # Titres
        fields = [(field.name, str(field.verbose_name),)
                  for field in chain(meta.fields, meta.many_to_many)
                  if field.name == code_field or not (field.auto_created or not (field.editable or self.non_editables))]
        for column, (field_code, field_name) in enumerate(fields, start=1):
            cell = worksheet.cell(row=1, column=column)
            cell.value = field_name
            cell.font = self.title_font
            column_letter = get_column_letter(column)
            widths[column_letter] = len(str(cell.value)) + CELL_OFFSET
        # Récupération des données
        queryset = model.objects.select_related().order_by(code_field)
        row = 2
        for element in queryset:
            for column, (field_code, field_name) in enumerate(fields, start=1):
                value = getattr(element, field_code)
                if value is None:
                    continue
                field = meta.get_field(field_code)
                if field.many_to_many:
                    m2m_code_field = getattr(field.related_model, '_code_field', 'id')
                    value = ', '.join(str(v) for v in value.values_list(m2m_code_field, flat=True))
                elif field.related_model is not None and field.related_model is MetaData:
                    if len(element.get_metadata()) > 0:
                        value = 'meta_{}_{}'.format(element._meta.model_name, row)
                        self.metadata[value] = []
                        for key_meta, value_meta in element.get_metadata().items():
                            self.metadata[value].append((key_meta, value_meta,))
                    else:
                        continue
                elif field.remote_field:
                    if not value:
                        value = ''
                    else:
                        value = getattr(value, code_field, value.id)
                elif field.choices:
                    value = getattr(element, 'get_{}_display'.format(field_code))()
                    if column not in dropdowns:
                        data_validation = dropdowns[column] = self.dropdowns[model, field_code]
                        worksheet.add_data_validation(data_validation)
                    dropdowns[column].add(worksheet["{}{}".format(get_column_letter(column), row)])
                elif field.get_internal_type() in ['DateField', 'DateTimeField']:
                    value = parsedate(value).isoformat()
                elif isinstance(value, FieldFile):
                    value = value.name
                elif isinstance(value, dict):
                    value = json_encode(value)
                cell = worksheet.cell(row=row, column=column)
                try:
                    cell.value = value
                except Exception:
                    cell.value = str(value)
                column_letter = get_column_letter(column)
                widths[column_letter] = max(widths[column_letter], len(str(value)) + CELL_OFFSET)
            row += 1
        # Ajout de lignes vides avec listes déroulantes
        for row in range(row, row + 10):
            for column, (field_code, field_name) in enumerate(fields, start=1):
                field = meta.get_field(field_code)
                if not field.choices:
                    continue
                if column not in dropdowns:
                    data_validation = dropdowns[column] = self.dropdowns[model, field_code]
                    worksheet.add_data_validation(data_validation)
                dropdowns[column].add(worksheet["{}{}".format(get_column_letter(column), row)])
        # Redimensionne les colonnes
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width
